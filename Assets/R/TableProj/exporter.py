#!/usr/bin/python
# -*- coding: UTF-8 -*-

#执行 pip install openpyxl 安装依赖

from sheet_data import SheetData
from sheet_data import get_cell_value
from openpyxl.cell.cell import Cell
import os
from openpyxl.descriptors import base
from openpyxl.descriptors.base import String
from openpyxl.reader.excel import load_workbook

# 比较文本，如果不同时写入文本
def write_text(path: str, text: str) -> bool:
    if os.path.exists(path):
        file = open(path, 'r', encoding='utf-8')
        if file.read() == text:
            return False
        file.close()
        os.remove(path)
    file = open(path, 'w', encoding='utf-8')
    file.write(text)
    file.close()

# 创建目录
def create_dir(dir: str):
    if not os.path.exists(dir):
        os.makedirs(dir)
# 目录处理
cmd = os.getcwd()
excel_dir = os.path.join(cmd, 'Excel')
create_dir(excel_dir)
json_dir = os.path.join(cmd, 'Json')
create_dir(json_dir)
client_data_dir = os.path.join(cmd, 'Client/Data')
create_dir(client_data_dir)
client_script_dir = os.path.join(cmd, 'Client/Script')
create_dir(client_script_dir)
server_data_dir = os.path.join(cmd, 'Server/Data')
create_dir(server_data_dir)
server_script_dir = os.path.join(cmd, 'Server/Script')
create_dir(server_script_dir)
language_dir = os.path.join(cmd, 'Language')
create_dir(language_dir)

# 表格中所有的 SheetData 数据
sheetdatas = list[list[str]]()

# 表格中需处理的语言
languages = dict[str, str]()

# 加载 excel 并 读取 sheet
for filename in os.listdir(excel_dir):
    # 处理打开 excel 时的缓存问题
    if filename.startswith('~$') or not filename.endswith('.xlsx'):
        continue
    excel_path = os.path.join(excel_dir, filename)
    excel_name = filename[0:filename.rfind('.')]
    try:
        wb = load_workbook(excel_path)
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            sheetdata = SheetData(excel_name, excel_path, sheetname)
            sheetdatas.append(sheetdata)
        wb.close()
    except Exception as e:
        raise RuntimeError('读取excel {0} 时出错，异常 {1}'.format(sheetname, e))
# 读取 sheet 并向 Language 表中添加语言项
for sheetdata in sheetdatas:
    try:
        sheetdata.read()
    except Exception as e:
        raise RuntimeError('读取excel {} 中的 sheet {0} 时出错，异常 {1}'.format(sheetdata.excel_name, sheetdata.sheet_name, e))
    else:
        if sheetdata.excel_name == '_Language':
            continue;
        for key in sheetdata.languages:
            languages[key] = sheetdata.languages[key]
        # 写入前后端数据
        client_data = sheetdata.to_cilent_data()
        client_script = sheetdata.to_cilent_script()
        server_data = sheetdata.to_server_data()
        server_script = sheetdata.to_server_script()
        json = sheetdata.to_json()
        if client_data:
            write_text(os.path.join(client_data_dir, sheetdata.sheet_name + '.txt'), client_data)
        if client_script:
            write_text(os.path.join(client_script_dir, 'DR' + sheetdata.sheet_name + '.cs'), client_script)
        if server_data:
            write_text(os.path.join(server_data_dir, sheetdata.sheet_name + '.txt'), server_data)
        if server_script:
            write_text(os.path.join(server_script_dir, sheetdata.sheet_name + '.txt'), server_script)
        if json:
            write_text(os.path.join(json_dir, sheetdata.sheet_name + '.json'), json)
        # 写入语言表
        lang_path = os.path.join(excel_dir, '_Language.xlsx')
        lang_wb = load_workbook(lang_path)
        lang_ws = lang_wb["Language"]
        for row_idx in range(6, lang_ws.max_row + 1):
            key = get_cell_value(lang_ws.cell(row_idx, 2))
            value = languages.get(key)
            if value:
                lang_ws.cell(row_idx, 5).value = value
                del languages[key]
        for key in languages:
            sys = key[0:key.find('_')]
            new_row_idx = None
            for row_idx in range(6, lang_ws.max_row + 1):
                other_key = get_cell_value(lang_ws.cell(row_idx, 2))
                other_sys = other_key[0:key.find('_')]
                if sys == other_sys:
                    new_row_idx = row_idx + 1
                    break
            if not new_row_idx:
                new_row_idx = lang_ws.max_row + 1
            lang_ws.insert_rows(new_row_idx)
            lang_ws.cell(new_row_idx, 2).value = key
            lang_ws.cell(new_row_idx, 5).value = languages[key]
        lang_wb.save(lang_path)
    
lang_suffix = {"ChineseSimplified": "zh-CN", "ChineseTraditional": "zh-TW", "English": "en", "Korean": "ko", }
lang_default = "ChineseSimplified"
# 导出语言表
for sheetdata in sheetdatas:
    if sheetdata.excel_name != '_Language':
        continue;
    langs = sheetdata.to_lang()
    for key in langs:
        xml_str = langs[key].decode('utf-8')
        suffix = lang_suffix.get(key)
        if suffix:
            write_text(os.path.join(language_dir, '{0}_{1}.xml'.format(sheetdata.sheet_name, suffix)), xml_str)
        else:
            raise RuntimeError('未发现语言后缀：{0}'.format(key))
        if key == lang_default:
            write_text(os.path.join(language_dir, '{0}.xml'.format(sheetdata.sheet_name)), xml_str)
print("EXPORT EXCELS DONE!!!")