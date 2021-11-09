#!/usr/bin/python
# -*- coding: UTF-8 -*-

import json
from os import error, path
import os
from variable_type import ArrayType, BoolType, DoubleType, FloatType, I18NType, IntType, LongType, ObjectType, StringType, VariableType
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import xml.dom.minidom as XmlDocument

# 读取 excel 中 cell 的文本值
def get_cell_value(cell: Cell):
    if cell and cell.value:
        return str(cell.value).strip()
    return '';

# 读取前端脚本模板
file = open('client_script_template.txt', encoding='utf-8')
client_script_template = file.read();
file.close()

# 可用的类型处理类
global_var_types = list[VariableType]()
global_var_types.append(BoolType())
global_var_types.append(StringType())
global_var_types.append(IntType())
global_var_types.append(LongType())
global_var_types.append(FloatType())
global_var_types.append(DoubleType())
global_var_types.append(ArrayType())
global_var_types.append(ObjectType())
global_var_types.append(I18NType())

# Sheet 数据类，提供读取和数据转换的功能
class SheetData:

    def __init__(self, excel_name: str, excel_path: str, sheet_name):
        self.excel_name = excel_name
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.comment = ""
        self.pro_col_idxs = list[int]()
        self.pro_comments = list[str]()
        self.pro_ports = list[str]()
        self.pro_types = list[str]()
        self.pro_names = list[str]()
        self.pro_datas = list[str]()
        self.var_types = list[str]()
        self.languages = dict[str, str]()

    # 读取 excel sheet的数据 
    def read(self) -> None:
        wb = load_workbook(self.excel_path)
        ws = wb[self.sheet_name]
        self.comment = get_cell_value(ws.cell(1, 2))
        # 读取 变量名称
        for cell in ws[5]:
            val = get_cell_value(cell)
            # 跳过注释
            if val == '' or val == '#':
                continue
            self.pro_col_idxs.append(cell.col_idx)
            self.pro_names.append(get_cell_value(cell))
        # 读取 变量其他数据
        for col_idx in self.pro_col_idxs:
            self.pro_comments.append(get_cell_value(ws.cell(2, col_idx)))
            self.pro_ports.append(get_cell_value(ws.cell(3, col_idx)))
            pro_type = get_cell_value(ws.cell(4, col_idx))
            self.pro_types.append(pro_type)
            filters = list(filter(lambda x: x.check(pro_type), global_var_types))
            if len(filters) == 1:
                self.var_types.append(filters[0])
            else:
                raise RuntimeError('为发现类型 {0} 对应的处理类'.format(pro_type))
        # 读取数据
        for row_idx in range(6, ws.max_row + 1):
            id = get_cell_value(ws.cell(row_idx, 2))
            flag = ws.cell(row_idx, 1).value
            # 跳过注释
            if id != '' and flag != '#':
                prop_data = []
                for idx in range(0, len(self.pro_col_idxs)):
                    col_idx = self.pro_col_idxs[idx]
                    val = get_cell_value(ws.cell(row_idx, col_idx))
                    var_type = self.var_types[idx]
                    pro_name = self.pro_names[idx]
                    # 未发现数据时启用默认
                    if not val:
                        val = var_type.default()
                    # 多语言会以 {表单名}_{变量名}_{ID} 为键插入语言表中
                    if var_type.check('i18n'):
                        key = '{0}_{1}_{2}'.format(self.sheet_name, pro_name, id)
                        self.languages[key] = val;
                        val = key;
                    prop_data.append(val)
                self.pro_datas.append(prop_data)
        wb.close();
    
    # 转换 sheet 为客户端数据
    def to_cilent_data(self) -> str:
        # 挑选端口
        if 'Id' in self.pro_names:
            idx = self.pro_names.index('Id')
            port = self.pro_ports[idx]
            if 'c' in port:
                rows = []
                for pro_data in self.pro_datas:
                    rows.append('\t'.join(pro_data))
                return '\n'.join(rows)
        return ''
    
    # 转换 sheet 为客户端脚本
    def to_cilent_script(self) -> str:
        prop_defines = ''
        prop_parses = ''
        for index in range(0, len(self.pro_col_idxs)):
            var_type = self.var_types[index]
            pro_comment = self.pro_comments[index]
            pro_port = self.pro_ports[index]
            pro_type = self.pro_types[index]
            pro_name = self.pro_names[index]
            # 挑选端口
            if 'c' in pro_port:
                prop_define, prop_parse = var_type.to_client_script(pro_type, pro_name, pro_comment)
                if pro_name != 'Id':
                    prop_defines += prop_define
                prop_parses += prop_parse
        return client_script_template.replace('__TYPE_COMMENT__', self.comment).replace('__TYPE_NAME__', self.sheet_name).replace('__PROP_DEFINE__', prop_defines).replace('__PROP_PARSE__', prop_parses)

    # 转换 sheet 为服务端数据
    def to_server_data(self) -> str:
        return '-------------'

    # 转换 sheet 为服务端脚本
    def to_server_script(self) -> str:
        return '-------------'

    # 转换 sheet 为 JSON
    def to_json(self) -> str:
        list = []
        for row_data in self.pro_datas:
            json_obj = {}
            for index in range(0, len(self.pro_col_idxs)):
                var_type = self.var_types[index]
                pro_comment = self.pro_comments[index]
                pro_port = self.pro_ports[index]
                pro_type = self.pro_types[index]
                pro_name = self.pro_names[index]
                json_obj[pro_name] = var_type.to_json(row_data[index])
            list.append(json_obj)
        return json.dumps({'list': list})

    # 转换为语言xml
    def to_lang(self) -> dict[str, str]:
        langs = {}
        for idx in range(0, len(self.pro_names)):
            pro_name = self.pro_names[idx]
            if pro_name == 'Key':
                continue
            doc = XmlDocument.Document()
            Dictionaries = doc.createElement('Dictionaries')
            doc.appendChild(Dictionaries)
            Dictionary = doc.createElement('Dictionary')
            Dictionary.attributes['Language'] = pro_name
            Dictionaries.appendChild(Dictionary)
            for pro_data in self.pro_datas:
                S = doc.createElement('String')
                Dictionary.appendChild(S)
                S.attributes['Key'] = pro_data[0]
                S.attributes['Value'] = pro_data[idx]
            langs[pro_name] = doc.toprettyxml(encoding='utf-8')
        return langs
